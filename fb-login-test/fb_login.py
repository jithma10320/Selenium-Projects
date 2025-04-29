from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.chrome.service import Service
import openpyxl
import time

# Load workbook and sheet
input_wb = openpyxl.load_workbook("C:/Users/User/Desktop/QA Projects/fb-login-with-selenium/fb_login.xlsx")
input_sheet = input_wb.active

# Create new workbook for results
output_wb = openpyxl.Workbook()
output_sheet = output_wb.active
output_sheet.append(["Username", "Password", "ExpectedResult", "ReadUsername", "ReadPassword", "ActualResult", "TestStatus"])

# ChromeDriver path setup
driver_path = "C:/Users/User/Desktop/QA Projects/fb-login-with-selenium/chromedriver-win64/chromedriver.exe"
service = Service(driver_path)
driver = webdriver.Chrome(service=service)

# Loop through test cases
for row in input_sheet.iter_rows(min_row=2, values_only=True):
    username = str(row[0]).strip()
    password = str(row[1]).strip()
    expected_result = str(row[2]).strip().lower()  # should be 'pass' or 'fail'

    driver.get("https://www.facebook.com/")
    time.sleep(2)

    try:
        # Clear and input credentials
        email_field = driver.find_element(By.ID, "email")
        pass_field = driver.find_element(By.ID, "pass")
        email_field.clear()
        pass_field.clear()
        email_field.send_keys(username)
        pass_field.send_keys(password)
        driver.find_element(By.NAME, "login").click()
        time.sleep(5)

        # Check login success via element (e.g., Home icon)
        try:
            driver.find_element(By.XPATH, '//a[@aria-label="Home"]')
            actual_result = "pass"
        except NoSuchElementException:
            actual_result = "fail"

        # Compare actual vs expected
        test_status = "Pass" if actual_result == expected_result else "Fail"

        # Append to results sheet
        output_sheet.append([username, password, expected_result, username, password, actual_result, test_status])

    except Exception as e:
        output_sheet.append([username, password, expected_result, username, password, "error", f"Error: {str(e)}"])

# Save and close
output_wb.save("C:/Users/User/Desktop/QA Projects/fb-login-with-selenium/login_test_results.xlsx")
driver.quit()